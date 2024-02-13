import json
import openpyxl
from datetime import datetime
# Path to the Excel file and the JSON output file
xlsx_file = 'data_purchase_order.xlsx'
json_output_file = 'output.json'

# Load data from Excel file
workbook = openpyxl.load_workbook(xlsx_file)
sheet = workbook.active

# Extract data from Excel sheet
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_dict = {}
    for idx, value in enumerate(row):
        column_name = sheet.cell(row=1, column=idx + 1).value
        if isinstance(value, datetime):
            value = value.isoformat()  # Convert datetime to ISO 8601 formatted string
        row_dict[column_name] = value
    data.append(row_dict)

# Generate JSON structure
json_data = {
    "data_grids": [
        {
            "content_id": "data_grid_21e8a4eceb61407b82cb1acf20d90e4f",
            "data": []
        }
    ]
}

for idx, row in enumerate(data):
    data_items = []
    for column_name, value in row.items():
        data_items.append({
            "data_item_id": column_name,
            "o": [{"v": value, "d": value}],
            "m": [{"v": value, "d": value}]
        })
    
    json_data["data_grids"][0]["data"].append({
        "id": idx,
        "data_items": data_items
    })

# Write JSON data to a file
with open(json_output_file, 'w') as json_file:
    json.dump(json_data, json_file, indent=4)

print("JSON data has been generated and saved to:", json_output_file)
